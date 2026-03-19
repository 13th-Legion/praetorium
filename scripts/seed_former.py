#!/usr/bin/env python3
"""Import inactive and blacklisted members from the roster XLSX."""

import openpyxl
import psycopg2
import re
from datetime import datetime

import os
import sys
DB_URL = os.environ.get("DATABASE_URL_SYNC")
if not DB_URL:
    print("ERROR: DATABASE_URL_SYNC environment variable is not set.", file=sys.stderr)
    sys.exit(1)
m = re.match(r'postgresql\+psycopg2://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', DB_URL)
DB_PARAMS = {
    "user": m.group(1), "password": m.group(2),
    "host": m.group(3), "port": int(m.group(4)), "dbname": m.group(5),
}

def parse_dt(val):
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()[:10]
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except:
        return None

def guess_nc(first, last):
    if not first or not last:
        return None
    return f"{first.strip().lower()}.{last.strip().lower()}"

def main():
    wb = openpyxl.load_workbook('/tmp/roster.xlsx')
    conn = psycopg2.connect(**DB_PARAMS)
    cur = conn.cursor()
    
    imported = 0
    
    # INACTIVE
    print("=== INACTIVE ===")
    ws = wb['Inactive']
    for row in ws.iter_rows(min_row=3, values_only=True):
        last = str(row[0] or '').strip()
        first = str(row[1] or '').strip()
        if not last or last in ('Former Members', 'Last', ''):
            continue
        
        callsign = str(row[2] or '').strip() or None
        phone = str(row[3] or '').strip() or None
        email = str(row[4] or '').strip() or None
        address = str(row[5] or '').strip() or None
        join_date = parse_dt(row[6])
        leave_date = parse_dt(row[7])
        rank = str(row[8] or '').strip() or None
        reason = str(row[9] or '').strip() or None
        nc_username = guess_nc(first, last)
        
        # Check for duplicate nc_username
        cur.execute("SELECT id FROM members WHERE nc_username = %s", (nc_username,))
        if cur.fetchone():
            for i in range(2, 10):
                alt = f"{nc_username}{i}"
                cur.execute("SELECT id FROM members WHERE nc_username = %s", (alt,))
                if not cur.fetchone():
                    nc_username = alt
                    break
        
        cur.execute("""
            INSERT INTO members (nc_username, first_name, last_name, callsign, email, phone,
                address, join_date, separation_date, status, company, state,
                primary_billet, is_veteran, has_ltc, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'inactive', '13th Legion', 'TX', %s, false, false, NOW(), NOW())
        """, (nc_username, first, last, callsign, email, phone, address, join_date, leave_date, reason))
        imported += 1
        print(f"  {last}, {first} -> inactive")
    
    # BLACKLIST
    print("\n=== BLACKLIST ===")
    ws = wb['Blacklist']
    for row in ws.iter_rows(min_row=3, values_only=True):
        last = str(row[0] or '').strip()
        first = str(row[1] or '').strip()
        if not last or last in ('Blacklisted People', 'Last', ''):
            continue
        
        callsign = str(row[2] or '').strip() or None
        phone = str(row[3] or '').strip() or None
        email = str(row[4] or '').strip() or None
        reason = str(row[5] or '').strip() or None
        nc_username = guess_nc(first, last)
        
        cur.execute("SELECT id FROM members WHERE nc_username = %s", (nc_username,))
        if cur.fetchone():
            for i in range(2, 10):
                alt = f"{nc_username}{i}"
                cur.execute("SELECT id FROM members WHERE nc_username = %s", (alt,))
                if not cur.fetchone():
                    nc_username = alt
                    break
        
        cur.execute("""
            INSERT INTO members (nc_username, first_name, last_name, callsign, email, phone,
                status, company, state, primary_billet, is_veteran, has_ltc, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, 'blacklisted', '13th Legion', 'TX', %s, false, false, NOW(), NOW())
        """, (nc_username, first, last, callsign, email, phone, reason))
        imported += 1
        print(f"  {last}, {first} -> blacklisted")
    
    conn.commit()
    
    # Reassign serial numbers to ALL members by join_date seniority
    print("\n=== REASSIGNING SERIAL NUMBERS ===")
    cur.execute("""
        WITH ranked AS (
            SELECT id, 
                   ROW_NUMBER() OVER (ORDER BY join_date ASC NULLS LAST, last_name ASC) as seq
            FROM members
        )
        UPDATE members SET 
            serial_seq = ranked.seq,
            serial_number = 'XIII-' || LPAD(ranked.seq::text, 4, '0')
        FROM ranked WHERE members.id = ranked.id;
    """)
    
    cur.execute("SELECT COUNT(*), COUNT(*) FILTER (WHERE status='inactive'), COUNT(*) FILTER (WHERE status='blacklisted') FROM members")
    total, inactive, blacklisted = cur.fetchone()
    
    cur.execute("SELECT serial_number, last_name, status FROM members WHERE last_name = 'Kavadas'")
    cav = cur.fetchone()
    
    conn.commit()
    cur.close()
    conn.close()
    
    print(f"\nImported {imported} former members")
    print(f"Total: {total} | Active+Recruit: {total-inactive-blacklisted} | Inactive: {inactive} | Blacklisted: {blacklisted}")
    print(f"Cav: {cav}")


if __name__ == "__main__":
    main()
